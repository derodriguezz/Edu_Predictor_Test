import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime
import unicodedata

### Autor: Jersson Quintero  Tel:3004069071

def Archivo_Registro_calificado(key):
    uploaded_file = st.file_uploader(f"Cargar archivo de {key}.xlsx", type=["xlsx", "xls"], key=key)
    if uploaded_file is not None:
        try:
            # Cargar el archivo y seleccionar la primera hoja
            libro = load_workbook(filename=uploaded_file, data_only=True)
            hoja = libro[libro.sheetnames[0]]
            # Extraer todos los datos de la hoja y crear el DataFrame
            datos = []
            for fila in hoja.iter_rows(values_only=True):
                datos.append(fila)

            df = pd.DataFrame(datos[1:], columns=datos[0])  # Usa la primera fila como encabezado

            # Eliminar espacios en los nombres de las columnas y en los datos de las celdas de texto
            df.columns = df.columns.map(str)  # Convertir nombres de columnas a string
            df.columns = df.columns.str.strip()
            df["RESOLUCION AMPARA"] = df["RESOLUCION AMPARA"].astype(str)
            df["RESOLUCION AMPARA FECHA"] = df["RESOLUCION AMPARA FECHA"].astype(str)
            df["MUNICIPIO"] = df["MUNICIPIO"].astype(str)
            df1=df.copy()
            for col in df.select_dtypes(include="object").columns:
                df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
            
            # # Formateo para columnas de enteros y decimales
            int_columns = df.select_dtypes(include='int').columns
            for col in int_columns:
                df[col] = df[col].apply(lambda x: f"{int(x):d}" if pd.notnull(x) else x)

            float_columns = df.select_dtypes(include='float').columns
            df[float_columns] = df[float_columns].apply(lambda x: f"{x:.2f}")
            for col in float_columns:
                df[col] = df[col].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else x)
            # Mostrar la vista previa del DataFrame
            st.write("Vista previa de los datos cargados:")
            st.dataframe(df.head())
            return df1
        except Exception as e:
            st.error(f"Error al leer el archivo: {e}. Por favor, revisarlo antes de cargar.")
    else:
        st.info("Por favor, sube un archivo Excel para continuar.")


def Cargar_archivos(key):
    uploaded_file1 = st.file_uploader(f"Cargar archivo de {key}.xlsx", type=["xlsx", "xls"],
                                      key=key)
    if uploaded_file1 is not None:
        try:
            # Intentar leer el archivo Excel usando pandas
            df = pd.read_excel(uploaded_file1, engine="openpyxl")
            
            # Mostrar las primeras filas del DataFrame
            st.write("Vista previa de los da    tos cargados:")
            # Formato personalizado: Sin comas para los enteros, decimales sin separadores de miles
            if key=='Metas_Proximo_Año':
                df["META_CUPOS"] = df["META_CUPOS"].astype(int)
            # Seleccionar solo las columnas con datos enteros
            df1=df.copy()
            int_columns = df.select_dtypes(include='int').columns
            
            # Aplicar el formato para eliminar las comas de las columnas de enteros
            for col in int_columns:
                df[col] = df[col].apply(lambda x: f"{int(x):d}" if pd.notnull(x) else x)
            
            # float_columns = df.select_dtypes(include='float64').columns
            # for col in float_columns:
            #     df[col] = df[col].apply(lambda x: f"{x:.2f}")
            st.dataframe(df.head())
            return df1
        except Exception as e:
            st.error(f"Error al leer el arc hivo {e}, por favor revisarlo antes de cargar")
            print("esta entrando allllll errorrrrrrr")
            
    else:
        st.info("Por favor, sube un archivo Excel para continuar.")


def calcular_proporcion_matriculados(df):
    #numero de matriculados en ese programa y esa modalidad entre el numero de años ofertados
    df['TOTAL_MATRICULADOS'] = pd.to_numeric(df['TOTAL_MATRICULADOS'], errors='coerce')
    df['AÑO_OFERTADOS'] = pd.to_numeric(df['AÑO_OFERTADOS'], errors='coerce')
    df['PROMEDIO_MATRICULADOS'] = (df['TOTAL_MATRICULADOS'] / df['AÑO_OFERTADOS']).fillna(0).round(3)
    df['SUM_CENTRO_MODALIDAD'] = df.groupby(['CODIGO_CENTRO', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION'])['PROMEDIO_MATRICULADOS'].transform('sum')
    df['PROMEDIO_CENTRO_MODALIDAD'] = (df['PROMEDIO_MATRICULADOS'] / df['SUM_CENTRO_MODALIDAD']).fillna(0).round(3)
    return df


def calcular_cupos_programa(df):
    # Calcular 'CUPOS_PROGRAMA'
    df['CUPOS_PROGRAMA'] = (df['PROMEDIO_CENTRO_MODALIDAD'] * df['META_CUPOS']).fillna(0)

    # Redondear los valores de 'CUPOS_PROGRAMA' al entero más cercano
    df['CUPOS_PROGRAMA'] = np.floor(df['CUPOS_PROGRAMA'] + 0.5).astype(int)
    df.loc[df['CUPOS_PROGRAMA'] < 0, 'CUPOS_PROGRAMA'] = 0
    return df

## AJUSTAR EL NUMERO DE CUPOS POR CENTRO DE FORMACIÓN
def ajustar_cupos_centro(df):
    df_prueba = df.copy()

    # Crear columna ID
    df_prueba['ID'] = df_prueba['CODIGO_CENTRO'].astype(str) + "_" + df_prueba['NIVEL_FORMACION'].astype(str) + "_" + df_prueba['MODALIDAD_FORMACION'].astype(str)
    # Sumar los valores por grupo ID
    df_prueba['suma_actual'] = df_prueba.groupby('ID')['CUPOS_PROGRAMA'].transform('sum')
    # Obtener el valor objetivo único por grupo
    df_prueba['valor_objetivo'] = df_prueba.groupby('ID')['META_CUPOS'].transform('first')
    # Calcular el factor de ajuste
    df_prueba['factor_ajuste'] = (df_prueba['valor_objetivo'] / df_prueba['suma_actual']).fillna(0).replace([np.inf, -np.inf], 0)
    # Ajustar los valores de 'valor' con el factor de ajuste
    df_prueba['valor_ajustado'] = df_prueba['CUPOS_PROGRAMA'] * df_prueba['factor_ajuste']
    # Redondear los valores ajustados al entero más cercano
    df_prueba['valor_ajustado'] = np.floor(df_prueba['valor_ajustado'] + 0.5).astype(int)
    # Verificar la diferencia entre la suma ajustada y el objetivo para cada grupo de ID
    df_prueba['diferencia'] = df_prueba.groupby('ID')['valor_ajustado'].transform('sum') - df_prueba['valor_objetivo']

    # Corregir la diferencia distribuyendo el excedente o déficit
    for id_grupo in df_prueba['ID'].unique():
        grupo = df_prueba[df_prueba['ID'] == id_grupo]
        diferencia_total = grupo['diferencia'].iloc[0]

        if diferencia_total != 0:
            # Reorganizar los grupos descontando matriculados de los programas con mayor o menor número de matriculados
            if diferencia_total > 0:
                grupo = grupo.sort_values(by='CUPOS_PROGRAMA', ascending=False)
            else:
                grupo = grupo.sort_values(by='CUPOS_PROGRAMA', ascending=True)

            # Seleccionar índices de los valores a ajustar
            indices_ajuste = grupo.index[:abs(int(diferencia_total))]

            # Ajustar sumando/restando 1 a los primeros valores para corregir la diferencia
            df_prueba.loc[indices_ajuste, 'valor_ajustado'] -= np.sign(diferencia_total)

    # Verificar que las sumas ajustadas ahora coincidan con los objetivos
    df_agrupado = df_prueba.groupby('ID')['valor_ajustado'].sum().reset_index()
    #df_agrupado
    df_prueba=df_prueba.drop(columns=['ID','suma_actual', 'valor_objetivo', 'factor_ajuste', 'diferencia'])
    df_prueba['CUPOS_PROGRAMA'] =df_prueba['valor_ajustado'].fillna(0).astype(int)
    df_prueba.drop(columns=['valor_ajustado'], inplace=True)
    return df_prueba

def calcular_numero_cursos(cupos, modalidad):
    # Definir los límites por modalidad
    if modalidad == 'P':  # Presencial
        max_cupos = 30
        min_cupos = 25
    elif modalidad == 'V':  # Virtual
        max_cupos = 50
        min_cupos = 30
    else:
        return None  # En caso de modalidad desconocida

    # Calcular el número de cursos
    if cupos >= min_cupos:
        # Usamos max_cupos para calcular el número máximo de cursos posibles
        numero_cursos = cupos // max_cupos
        # Si sobran cupos que permitan crear un curso adicional con al menos el número mínimo de cupos
        if cupos % max_cupos >= min_cupos:
            numero_cursos += 1
    else:
        # Si los cupos no alcanzan para al menos el mínimo de un curso
        numero_cursos = 0

    return numero_cursos


def procesar_dataframe(df):
        # Creamos una lista para almacenar los DataFrames procesados
        dfs = []

        # Iteramos por cada combinación única de CODIGO_CENTRO, NIVEL_FORMACION, MODALIDAD_FORMACION
        for (centro, nivel, modalidad), group in df.groupby(['CODIGO_CENTRO', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION']):
            # Ordenamos por NUMERO_FICHAS de menor a mayor
            group = group.sort_values(by='NUMERO_CURSOS')

            # Filtramos aquellos que tienen NUMERO_FICHAS = 0
            fichas_cero = group[group['NUMERO_CURSOS'] == 0]

            if not fichas_cero.empty:
                # Ordenamos por CUPOS_PROGRAMA de menor a mayor
                fichas_cero = fichas_cero.sort_values(by='CUPOS_PROGRAMA')

                # Eliminamos el primer elemento con NUMERO_FICHAS = 0
                fichas_cero = fichas_cero.iloc[1:]

                # Combinamos los datos modificados (sin el primer registro) con el resto del grupo
                group = pd.concat([group[group['NUMERO_CURSOS'] != 0], fichas_cero])

            # Agregamos el grupo procesado a la lista
            dfs.append(group)

        # Concatenamos todos los grupos procesados en un único DataFrame
        df_final = pd.concat(dfs)

        return df_final


# prompt: hacer un for por cada codigo_centro, nivel:formacion, modalidad_formacion dataframe df_union_nombre_programa  para sumar por PROGRAMA_FORMACION iguales las columnas %_CERTIFICADOS, CUPOS_PROGRAMA, NUMERO_FICHAS, y en el cacso de %_certificados, hacer una funcion que si son dos numeros sumarlos, si es un numero y un desconocido colocar el numero, si solo son desconocimiento asignar desconocimiento

def procesar_certificados(valores_certificados):
    """Procesa una lista de valores de certificados para obtener un valor final."""
    if all(valor == 'Desconocimiento' for valor in valores_certificados):
        return 'Desconocimiento'
    elif 'Desconocimiento' in valores_certificados:
        valores_numericos = [valor for valor in valores_certificados if valor != 'Desconocimiento']
        if valores_numericos:
            return sum(valores_numericos)/len(valores_numericos)
        else:
            return 'Desconocimiento'
    else:
        return sum(valores_certificados)
        

def eliminar_tildes_y_puntos(texto):
    texto = texto.replace('.', '')  # Eliminar puntos
    texto = texto.strip()  # Eliminar espacios finales
    texto = ''.join((c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn'))
    #print(texto)
    return texto

def mostrar_estimacion_fichas():
    st.title("Estimación de Fichas")
    st.write("Aquí puedes agregar contenido específico para la sección de estimación de fichas.")

    # Lectura de los archivos .xlsx
    df_entrada = Cargar_archivos(key="Base_Historica")
    df_leer_vigencias = Archivo_Registro_calificado(key="Registro_Calificado")
    df_metas = Cargar_archivos(key="Metas_Proximo_Año")

    if st.button("Calcular CUPOS_PROXIMO_AÑO"):
        nuevo_df = Calcular_Cupos(df_entrada,df_leer_vigencias,df_metas)
        st.write("Nuevo DataFrame calculado:")
        st.dataframe(nuevo_df)

        # Botón para descargar el nuevo DataFrame como Excel
        def descargar_excel(dataframe):
            output = BytesIO()
            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                dataframe.to_excel(writer, index=False, sheet_name="Hoja1")
            output.seek(0)
            return output

        excel_data = descargar_excel(nuevo_df)
        st.download_button(
            label="Descargar como Excel",
            data=excel_data,
            file_name="nuevo_dataframe.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

############################### desarrollo del algoritmo de estimación de fichas para los proximos años
def Calcular_Cupos(df_entrada,df_leer_vigencias,df_metas):
    df_copia = df_entrada.copy()

    #limpieza de columas que no generan valor
    #eliminar las columas FICHA, CODIGO_REGION,NOMBRE_CENTRO,VERSION_PROGRAMA,CODIGO_VERSION,TIPO_RESPUESTA,POF_ID,TIPO_OFERTA  del dataframedf_copia
    columnas_a_eliminar = ['FICHA', 'CODIGO_REGIONAL','NOMBRE_REGIONAL','NOMBRE_CENTRO','VERSION_PROGRAMA','CODIGO_VERSION','TIPO_RESPUESTA','POF_ID','TIPO_OFERTA']
    df_copia = df_copia.drop(columns=columnas_a_eliminar)

    ##FILTRAR LOS PROGRAMAS QUE ESTAN EN EJECUCIÓN Y LOS QUE ESTAN DEMANDADOS DESDE EL 2022 A LA ACTUALIDAD
    # filtra los programas que estan en ejecución
    agrupacion_ejecucion = df_copia[df_copia['ESTADO_PROGRAMA'] == 'En ejecucion'].copy()

    years_list = df_copia['AÑO_FICHA'].unique().tolist()
    years_actuales=years_list[-3:]
    years_anteriores=years_list[:-3]
    years_actuales

    # Comparación de los
    programas_anteriores = agrupacion_ejecucion[agrupacion_ejecucion['AÑO_FICHA'].isin(years_anteriores)].groupby(
        ['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'MODALIDAD_FORMACION'])[['MATRICULADOS', 'INSCRITOS']].sum().reset_index()

    programas_actuales = agrupacion_ejecucion[agrupacion_ejecucion['AÑO_FICHA'].isin(years_actuales)].groupby(
        ['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'MODALIDAD_FORMACION'])[['MATRICULADOS', 'INSCRITOS']].sum().reset_index()

    # Crear una combinación clave de 'DATOS_CENTRO' y 'CODIGO_PROGRAMA_FORMACION' en ambos subconjuntos
    programas_anteriores['COMBINACION'] = programas_anteriores['CODIGO_CENTRO'].astype(str) + "_" + programas_anteriores['CODIGO_PROGRAMA_FORMACION'].astype(str)+"_" + programas_anteriores['MODALIDAD_FORMACION'].astype(str)
    programas_actuales['COMBINACION'] = programas_actuales['CODIGO_CENTRO'].astype(str) + "_" + programas_actuales['CODIGO_PROGRAMA_FORMACION'].astype(str)+"_" + programas_actuales['MODALIDAD_FORMACION'].astype(str)

    # Verificar si las combinaciones de 2019-2021 están en 2022-2024
    programas_anteriores['CONTINUIDAD'] = programas_anteriores['COMBINACION'].isin(programas_actuales['COMBINACION']).map({True: 'SI', False: 'NO'})

    # Eliminar la columna 'COMBINACION' si ya no es necesaria
    programas_anteriores = programas_anteriores.drop(columns=['COMBINACION'])
    programas_actuales = programas_actuales.drop(columns=['COMBINACION'])
    # 1. Filtrar los registros de 2019-2021 donde 'CONTINUIDAD' = 'SI'
    programas_anteriores_continuos = programas_anteriores[programas_anteriores['CONTINUIDAD'] == 'SI'].copy()

    # Crear el DataFrame programas_no_continuos
    programas_descontinuados = programas_anteriores[programas_anteriores['CONTINUIDAD'] == 'NO'].copy()

    # 2. Eliminar la columna 'CONTINUIDAD' de ambos DataFrames
    programas_anteriores_continuos = programas_anteriores_continuos.drop(columns=['CONTINUIDAD'])

    # 3. Unir los registros de 2019-2021 con 'SI' en CONTINUIDAD con los registros de 2022-2024
    programas_continuos = pd.concat([programas_anteriores_continuos, programas_actuales], ignore_index=True)

    #4_JA. Se unen los grupos que es repitieron en los grupos 2019-2021 y 2022-2024, sumando matriculados e inscritos
    programas_continuos = programas_continuos.groupby(
        ['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'MODALIDAD_FORMACION'])[['MATRICULADOS', 'INSCRITOS']].sum().reset_index()

    filtro_programas_actuales = agrupacion_ejecucion[agrupacion_ejecucion['CODIGO_PROGRAMA_FORMACION'].isin(programas_continuos['CODIGO_PROGRAMA_FORMACION'])]
    ##FILTRAR LOS PROGRAMAS QUE ESTAN EN EJECUCIÓN Y LOS QUE ESTAN DEMANDADOS DESDE EL 2022 A LA ACTUALIDAD
    # filtra los programas que estan en ejecución
    agrupacion_ejecucion = df_copia[df_copia['ESTADO_PROGRAMA'] == 'En ejecucion'].copy()

    # Comparación de los
    programas_anteriores = agrupacion_ejecucion[agrupacion_ejecucion['AÑO_FICHA'].isin(years_anteriores)].groupby(
        ['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'MODALIDAD_FORMACION'])[['MATRICULADOS', 'INSCRITOS']].sum().reset_index()

    programas_actuales = agrupacion_ejecucion[agrupacion_ejecucion['AÑO_FICHA'].isin(years_actuales)].groupby(
        ['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'MODALIDAD_FORMACION'])[['MATRICULADOS', 'INSCRITOS']].sum().reset_index()

    # Crear una combinación clave de 'DATOS_CENTRO' y 'CODIGO_PROGRAMA_FORMACION' en ambos subconjuntos
    programas_anteriores['COMBINACION'] = programas_anteriores['CODIGO_CENTRO'].astype(str) + "_" + programas_anteriores['CODIGO_PROGRAMA_FORMACION'].astype(str)+"_" + programas_anteriores['MODALIDAD_FORMACION'].astype(str)
    programas_actuales['COMBINACION'] = programas_actuales['CODIGO_CENTRO'].astype(str) + "_" + programas_actuales['CODIGO_PROGRAMA_FORMACION'].astype(str)+"_" + programas_actuales['MODALIDAD_FORMACION'].astype(str)

    # Verificar si las combinaciones de 2019-2021 están en 2022-2024
    programas_anteriores['CONTINUIDAD'] = programas_anteriores['COMBINACION'].isin(programas_actuales['COMBINACION']).map({True: 'SI', False: 'NO'})

    # Eliminar la columna 'COMBINACION' si ya no es necesaria
    programas_anteriores = programas_anteriores.drop(columns=['COMBINACION'])
    programas_actuales = programas_actuales.drop(columns=['COMBINACION'])
    # 1. Filtrar los registros de 2019-2021 donde 'CONTINUIDAD' = 'SI'
    programas_anteriores_continuos = programas_anteriores[programas_anteriores['CONTINUIDAD'] == 'SI'].copy()

    # Crear el DataFrame programas_no_continuos
    programas_descontinuados = programas_anteriores[programas_anteriores['CONTINUIDAD'] == 'NO'].copy()

    # 2. Eliminar la columna 'CONTINUIDAD' de ambos DataFrames
    programas_anteriores_continuos = programas_anteriores_continuos.drop(columns=['CONTINUIDAD'])

    # 3. Unir los registros de 2019-2021 con 'SI' en CONTINUIDAD con los registros de 2022-2024
    programas_continuos = pd.concat([programas_anteriores_continuos, programas_actuales], ignore_index=True)

    #4_JA. Se unen los grupos que es repitieron en los grupos 2019-2021 y 2022-2024, sumando matriculados e inscritos
    programas_continuos = programas_continuos.groupby(
        ['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'MODALIDAD_FORMACION'])[['MATRICULADOS', 'INSCRITOS']].sum().reset_index()

    filtro_programas_actuales = agrupacion_ejecucion[agrupacion_ejecucion['CODIGO_PROGRAMA_FORMACION'].isin(programas_continuos['CODIGO_PROGRAMA_FORMACION'])]
    # filtro_programas_actuales

    ##PERMITE CALCULAR LA PROPORCIÓN PARA LA POSTERIOR ASIGNACIÓN DE CUPOS
    #crear la agrupación de datos sumando el numero de matriculados, el valor unico de año ofertado por CODIGO_CENTRO, CODIGO_PROGRAMA_FORMACION y NIVEL_FORMACION, MODALIDAD_FORMACION
    #agrupacion_ejecucion = filtro_2022_2024[filtro_2022_2024['ESTADO_PROGRAMA'] == 'En ejecucion'].copy()

    agrupacion_ejecucion = filtro_programas_actuales.groupby(['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION','AÑO_FICHA'])['MATRICULADOS'].agg(['sum']).reset_index()

    agrupacion_ejecucion = agrupacion_ejecucion.groupby(['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION']).agg(
        TOTAL_MATRICULADOS=('sum', 'sum'),
        AÑO_OFERTADOS = ('AÑO_FICHA', 'count')).reset_index()

    # funcion que calcula el promedio y proporcion por centro de formación segun los programas que hay
    

    agrupacion_ejecucion = calcular_proporcion_matriculados(agrupacion_ejecucion)

    """25/10/2025 agregar por resoluciones vigentes"""


    # from openpyxl import load_workbook
    # Ruta al archivo y Cargar el archivo Excel


    df_vigencias = df_leer_vigencias.copy()
    columnas_a_eliminar = ['TIPO DE TRÁMITE','FECHA RADICADO','NUMERO DE RESOLUCION','FECHA DE RESOLUCION','RESUELVE',
                        'DECRETO QUE AMPARA','SNIES','NUMERO DE RESOLUCION','RED DE  CONOCIMIENTO','TIPO SEDE','MUNICIPIO',
                        'LUGAR DE DESARROLLO','DIRECCION','OBSERVACIONES','CLASIFICACIÓN PARA TRÁMITE','APRENDICES PRIMER COHORTE']
    df_vigencias = df_vigencias.drop(columns=columnas_a_eliminar)
    df_vigencias.info()

    # prompt: cambiar el nombre de las columnas Regional por CODIGO_REGIONAL, COD DEL PROGRAMA POR CODIGO_PROGRAMA_FORMACION CODIGO_CENTRO	CODIGO_PROGRAMA_FORMACION	NIVEL_FORMACION	MODALIDAD_FORMACION

    # Rename columns
    df_vigencias.rename(columns={
        'REGIONAL': 'CODIGO_REGIONAL',
        'COD DEL PROGRAMA': 'CODIGO_PROGRAMA_FORMACION',
        'CENTRO DE FORMACIÓN': 'CODIGO_CENTRO',  # This one was already correct
        'NIVEL DE FORMACIÓN': 'NIVEL_FORMACION',
        'MODALIDAD': 'MODALIDAD_FORMACION'}, inplace=True)
    df_vigencias.info()

    # prompt: reordenar las columnas de df_vigencias comenzando con estas columnas 'REGIONAL','CODIGO_CENTRO','CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION'

    # Define the desired order of columns
    new_column_order = ['PROCESO','CODIGO_REGIONAL','CODIGO_CENTRO','CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION']

    # Get the remaining columns not in the new order
    remaining_columns = [col for col in df_vigencias.columns if col not in new_column_order]

    # Combine the new order and remaining columns
    final_column_order = new_column_order + remaining_columns

    # Reorder the DataFrame columns
    df_vigencias = df_vigencias[final_column_order]
    df_vigencias


    df_vigencias_group=df_vigencias[(df_vigencias['VIGENCIA RC']  == 'RC Activo')]
    df_vigencias_group=df_vigencias_group.groupby(['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION','VIGENCIA RC']).size().reset_index(name='COUNT')
    df_vigencias_group.drop(columns=['COUNT'], inplace=True)
    df_vigencias_group

    agrupacion_ejecucion_Tecnologo = agrupacion_ejecucion[
        (agrupacion_ejecucion['NIVEL_FORMACION'] == 'TECNÓLOGO') |
        (agrupacion_ejecucion['NIVEL_FORMACION'] == 'ESPECIALIZACIÓN TECNOLÓGICA')].copy()

    #eliminar las
    agrupacion_ejecucion = agrupacion_ejecucion[~agrupacion_ejecucion.set_index(['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION']).index.isin(agrupacion_ejecucion_Tecnologo.set_index(['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION']).index)]
    agrupacion_ejecucion

    df_vigencias_group['NIVEL_FORMACION'] = df_vigencias_group['NIVEL_FORMACION'].replace({'TECNOLOGIA': 'TECNÓLOGO', 'ESPECIALIZACION TECNOLOGICA': 'ESPECIALIZACIÓN TECNOLÓGICA'})
    df_vigencias_group['MODALIDAD_FORMACION'] = df_vigencias_group['MODALIDAD_FORMACION'].replace({'PRESENCIAL': 'P', 'VIRTUAL': 'V','A DISTANCIA':'A'})
    agrupacion_ejecucion_Tecnologo

    # Seleccionar las columnas relevantes en df_vigencias para la comparación
    columnas_comunes = ['CODIGO_CENTRO','CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION','MODALIDAD_FORMACION']

    # Realizar el merge entre ambos DataFrames, para quedarte solo con los registros que coinciden
    agrupacion_ejecucion_filtrado = agrupacion_ejecucion_Tecnologo.merge(
        df_vigencias_group[columnas_comunes],
        on=columnas_comunes,
        how='inner'  # Usamos inner para obtener solo las filas que coinciden
    )

    # Concatenar los DataFrames
    agrupacion_ejecucion = pd.concat([agrupacion_ejecucion, agrupacion_ejecucion_filtrado], ignore_index=True)

    # Eliminar las filas de agrupacion_ejecucion_Tecnologo que están en agrupacion_ejecucion_filtrado
    agrupacion_ejecucion_Tecnologo_no_vigencia = agrupacion_ejecucion_Tecnologo[~agrupacion_ejecucion_Tecnologo.set_index(
        ['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION']).index.isin(
        agrupacion_ejecucion_filtrado.set_index(['CODIGO_CENTRO', 'CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION']).index)]

    agrupacion_ejecucion = calcular_proporcion_matriculados(agrupacion_ejecucion)

    #LEER EL DOCUMENTO METAS PARA AÑOS POSTERIORES
    #leer un archivo de excel .xlsx y convertirlo en un dataframe df
    ##LIMPIEZA Y NORMALIZACIÓN DE LOS DATOS

    df_metas.rename(columns={'MODALIDAD': 'MODALIDAD_FORMACION'}, inplace=True)
    df_metas['MODALIDAD_FORMACION'] = df_metas['MODALIDAD_FORMACION'].replace({'PRESENCIAL': 'P', 'VIRTUAL': 'V'})

    """Fin de agregar las resolciones vigentes
    """
    ##
    asociacion = pd.merge(
        agrupacion_ejecucion,  # DataFrame base
        df_metas[['CODIGO_CENTRO', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION','META_CUPOS']],  # Seleccionamos las columnas relevantes de df_metas
        on=['CODIGO_CENTRO', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION'],  # Claves de unión
        how='left'  # Mantener todas las filas de agrupacion_ejecucion
    )
    asociacion['META_CUPOS'] = asociacion['META_CUPOS'].fillna(0)
    df_calculo_cupos_2025 = asociacion.copy()

    df_calculo_cupos_2025 = calcular_cupos_programa(df_calculo_cupos_2025)

    ## EXPORTAR LOS PROGRAMAS QUE NO TIENES CUPOS PORQUE DADO QUE NO TIENEN CUPOS O QUE NO TIENEN DEMANDA O NO ESTAN EN EJECUCIÓN
    df_programas_cero_cupos=df_calculo_cupos_2025[df_calculo_cupos_2025['CUPOS_PROGRAMA'] == 0].copy()
    df_calculo_cupos_2025=df_calculo_cupos_2025[df_calculo_cupos_2025['CUPOS_PROGRAMA'] > 0]


    df_calculo_cupos_2025 = calcular_proporcion_matriculados(df_calculo_cupos_2025)
    df_calculo_cupos_2025 = calcular_cupos_programa(df_calculo_cupos_2025)


    # Llamada a la función
    df_calculo_cupos_2025 = ajustar_cupos_centro(df_calculo_cupos_2025)

    # Mostrar el DataFrame final ajustado

    # funcion que calcula el numero de cursos que puedan contener los cupos_2025,
    # estos cursos maximo deben ser de 30 y minimo 25 para la MODALIDAD_FORMACIÓN P
    #Y PARA LA MODALIDAD_FORMACION V DEBEN SER DE MAXIMO 50 MINIMO 30 EN EL DATAFRAME df_grupo_ponderado_cupos

    # Aplicar la función al DataFrame
    df_calculo_cupos_2025['NUMERO_CURSOS'] = df_calculo_cupos_2025.apply(
        lambda row: calcular_numero_cursos(row['CUPOS_PROGRAMA'], row['MODALIDAD_FORMACION']), axis=1
    )

    df_calculo_cupos_2025['NUMERO_CURSOS'] = df_calculo_cupos_2025['NUMERO_CURSOS'].fillna(0)
    # Mostrar el DataFrame con la nueva columna


    """# Realizar el ajuste de las fichas a ofertar"""

    # si el NUMERO_CURSOS es 0 copiar en un dataframe con nombre MATRICULADOS_NO_CURSOS Y eliminar del dataframe df_calculo_cupos_2025

    MATRICULADOS_POSIBLE_NO_CURSOS = df_calculo_cupos_2025[df_calculo_cupos_2025['NUMERO_CURSOS'] == 0].copy()

    df_calculo_cupos_2025_ajuste = df_calculo_cupos_2025.copy()


    # Llamamos a la función con tu DataFrame
    contador=0
    while df_calculo_cupos_2025_ajuste[df_calculo_cupos_2025_ajuste['NUMERO_CURSOS'] == 0].empty == False:
        df_calculo_cupos_2025_ajuste = procesar_dataframe(df_calculo_cupos_2025_ajuste)
        df_calculo_cupos_2025_ajuste = calcular_proporcion_matriculados(df_calculo_cupos_2025_ajuste)
        df_calculo_cupos_2025_ajuste = ajustar_cupos_centro(df_calculo_cupos_2025_ajuste)

        df_calculo_cupos_2025_ajuste['NUMERO_CURSOS'] = df_calculo_cupos_2025_ajuste.apply(
            lambda row: calcular_numero_cursos(row['CUPOS_PROGRAMA'], row['MODALIDAD_FORMACION']), axis=1 )
        df_calculo_cupos_2025_ajuste['NUMERO_CURSOS'] = df_calculo_cupos_2025_ajuste['NUMERO_CURSOS'].fillna(0)
        contador+=1
        print(contador)
        if contador > 100:
            break

    # df_calculo_cupos_2025_ajuste

    # resultado = df_calculo_cupos_2025.groupby(['CODIGO_CENTRO', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION'])
    # conteo_por_grupo = resultado.size()
    # print(conteo_por_grupo)
    # fichas_cero = df_calculo_cupos_2025[df_calculo_cupos_2025['NUMERO_CURSOS'] == 0]
    # fichas_cero

    """# PARTE DE CERTIFICADOS"""

    # obtener el dato del numero de certificados por programa de formación en ejecución y ficha de formación terminada.
    fichas_terminadas=filtro_programas_actuales[filtro_programas_actuales['ESTADO_FICHA'] == 'Terminada'].copy()
    programas_terminados=fichas_terminadas.groupby(['CODIGO_CENTRO','CODIGO_PROGRAMA_FORMACION','NIVEL_FORMACION','MODALIDAD_FORMACION']).agg(
        TOTAL_MATRICULADOS=('MATRICULADOS', 'sum'),
        TOTAL_RETIROS = ('RETIROS_VOLUNTARIOS', 'sum'),
        TOTAL_CANCELADOS = ('CANCELADOS', 'sum'),
        TOTAL_POR_CERTIFICAR = ('POR_CERTIFICAR', 'sum'),
        TOTAL_CERTIFICADOS = ('CERTIFICADOS', 'sum')
        ).reset_index()

    programas_terminados['%_CERTIFICADOS'] = (programas_terminados['TOTAL_CERTIFICADOS'] / programas_terminados['TOTAL_MATRICULADOS']).fillna(0).round(3) *100


    #union de los programas a ofertar con los cupos y las fichas con el calculo de los certificados
    resultados = pd.merge(
        df_calculo_cupos_2025_ajuste,  # DataFrame base
        programas_terminados[['CODIGO_CENTRO','CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION','%_CERTIFICADOS']],  # Seleccionamos las columnas relevantes de df_metas
        on=['CODIGO_CENTRO','CODIGO_PROGRAMA_FORMACION', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION'],  # Claves de unión
        how='left'  # Mantener todas las filas de agrupacion_ejecucion
    )
    resultados['%_CERTIFICADOS'] = resultados['%_CERTIFICADOS'].fillna('Desconocimiento')
    # resultados

    ##organizar el dataframe para obtener la salida esperada
    df_final=resultados.copy()
    df_copia1=df_copia.groupby(['CODIGO_CENTRO','NIVEL_FORMACION','CODIGO_PROGRAMA_FORMACION','MODALIDAD_FORMACION','DATOS_CENTRO','PROGRAMA_FORMACION']).size().reset_index(name='COUNT')
    # df_copia1

    df_final=pd.merge(df_final,df_copia1[['CODIGO_CENTRO','NIVEL_FORMACION','CODIGO_PROGRAMA_FORMACION','MODALIDAD_FORMACION','DATOS_CENTRO','PROGRAMA_FORMACION']],
                    on=['CODIGO_CENTRO', 'NIVEL_FORMACION','CODIGO_PROGRAMA_FORMACION', 'MODALIDAD_FORMACION'],
                    how='left')
    df_final.rename(columns={'NUMERO_CURSOS': 'NUMERO_FICHAS'}, inplace=True)
    #df_final=df_final[['CODIGO_CENTRO','DATOS_CENTRO','NIVEL_FORMACION','MODALIDAD_FORMACION','CODIGO_PROGRAMA_FORMACION','PROGRAMA_FORMACION','%_CERTIFICADOS','CUPOS_PROGRAMA','NUMERO_FICHAS']]

    # df_final

    # prompt:  cambiar las letras que tienen tilde y eliminar los puntos, y los espacios finales de la columna programa_formacion e imprimir los que se cambiaron

    df_union_nombre_programa=df_final.copy()

    cambios = {}
    for i, row in df_union_nombre_programa.iterrows():
        programa_original = row['PROGRAMA_FORMACION']
        programa_procesado = eliminar_tildes_y_puntos(programa_original)
        if programa_original != programa_procesado:
            cambios[programa_original] = programa_procesado
        df_union_nombre_programa.at[i, 'PROGRAMA_FORMACION'] = programa_procesado


    # print("Cambios realizados:", len(cambios))
    # for original, procesado in cambios.items():
    #   print(f"- {original} -> {procesado}")

    # df_union_nombre_programa

    df_agrupado_nombre_programas = pd.DataFrame()
    for (centro, nivel, modalidad), grupo in df_union_nombre_programa.groupby(['CODIGO_CENTRO', 'NIVEL_FORMACION', 'MODALIDAD_FORMACION']):
        # Agrupamos por programa de formación dentro del grupo
        for programa, sub_grupo in grupo.groupby(['PROGRAMA_FORMACION']) :
            # Sumamos %_CERTIFICADOS, CUPOS_PROGRAMA, NUMERO_FICHAS
            suma_certificados = procesar_certificados(sub_grupo['%_CERTIFICADOS'].tolist())
            suma_cupos = sub_grupo['CUPOS_PROGRAMA'].sum()
            suma_fichas = sub_grupo['NUMERO_FICHAS'].sum()

            # Creamos un nuevo registro con la información agregada
            nuevo_registro = pd.DataFrame({
                'CODIGO_CENTRO': [centro],
                'DATOS_CENTRO': [sub_grupo['DATOS_CENTRO'].iloc[0]],
                'NIVEL_FORMACION': [nivel],
                'MODALIDAD_FORMACION': [modalidad],
                'CODIGO_PROGRAMA_FORMACION': [sub_grupo['CODIGO_PROGRAMA_FORMACION'].iloc[0]],
                'PROGRAMA_FORMACION': programa,
                '%_CERTIFICADOS': [suma_certificados],
                'CUPOS_PROGRAMA': [suma_cupos],
                'NUMERO_FICHAS': [suma_fichas]
            })

            # Concatenamos el nuevo registro al DataFrame final
            df_agrupado_nombre_programas = pd.concat([df_agrupado_nombre_programas, nuevo_registro], ignore_index=True)

    # df_agrupado_nombre_programas

    # Aplicar la función al DataFrame
    df_agrupado_nombre_programas['NUMERO_FICHAS'] = df_agrupado_nombre_programas.apply(
        lambda row: calcular_numero_cursos(row['CUPOS_PROGRAMA'], row['MODALIDAD_FORMACION']), axis=1
    )

    # Mostrar el DataFrame con la nueva columna
    #df_agrupado_nombre_programas

    # # Exportar el DataFrame a un archivo Excel
    # df_final.to_excel('df_final.xlsx', index=False)

    # Convertir la fecha a string
    # Obtener la fecha del hoy
    fecha_actual = datetime.now()
    fecha_actual_str = fecha_actual.strftime("%Y%m%d")
    current_year=fecha_actual.strftime("%Y")
    next_year=int(current_year)+1
    df_agrupado_nombre_programas['AÑO_ESTIMACION']=next_year

    Nombre_salida=fecha_actual_str+"_Cupos_Ofertados_AÑO"+str(next_year)+".xlsx"
    #print(Nombre_salida)

    # # Exportar el DataFrame a un archivo Excel
    df_agrupado_nombre_programas.to_excel(Nombre_salida, index=False)
    return df_agrupado_nombre_programas


### Autor: Jersson Quintero  Tel:3004069071